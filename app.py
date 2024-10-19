from flask import Flask, render_template, request, redirect, url_for
import openpyxl
import os
from pyngrok import ngrok

app = Flask(__name__)

# Registration Route
@app.route("/", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        name = request.form["name"]
        address = request.form["address"]
        phone = request.form["phone"]
        email = request.form["email"]
        
        # Store user data in details.xlsx
        workbook = openpyxl.load_workbook("details.xlsx") 
        sheet = workbook.active
        sheet.append([name, address, phone, email])
        workbook.save("details.xlsx")
        
        # Redirect to attendance page with query parameters
        return redirect(url_for("attendance", name=name, phone=phone))
    
    return render_template("register.html")  # Make sure the template has the updated background image

# Attendance Route
@app.route("/attendance", methods=["GET", "POST"])
def attendance():
    if request.method == "POST":
        name = request.form["name"]
        phone = request.form["phone"]
        status = request.form["status"]
        
        if status == "Present":
            try:
                # Load the workbook
                workbook = openpyxl.load_workbook("attendance.xlsx")
                sheet = workbook.active
                
                # Append name and phone of present user
                sheet.append([name, phone])
                
                # Save the workbook
                workbook.save("attendance.xlsx")
                print(f"Attendance for {name} with phone {phone} has been recorded.")
                
            except Exception as e:
                print(f"Error handling attendance: {e}")
        
        return redirect(url_for("attendance_success"))
    
    name = request.args.get("name")
    phone = request.args.get("phone")
    
    return render_template("attendance.html", name=name, phone=phone)

# Success Page
@app.route("/attendance_success")
def attendance_success():
    return "Attendance recorded successfully!"

# Create Excel Files if Not Exist
def create_excel_files():
    if not os.path.exists("details.xlsx"):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Address", "Phone", "Email"])  # Add headers
        workbook.save("details.xlsx")

    if not os.path.exists("attendance.xlsx"):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Phone"])  # Only store present user's Name and Phone
        workbook.save("attendance.xlsx")

# Ensure Excel files exist
create_excel_files()

if __name__ == "__main__":
    app.run(debug=False,host='0.0.0.0')
    
    """   # Start ngrok tunnel
   public_url = ngrok.connect(5000)  # Change this to 5000
   print(f" * Ngrok tunnel available at: {public_url}")

   app.run(port=5000)  # Ensure this matches the ngrok port"""
